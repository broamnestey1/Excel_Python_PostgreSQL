import openpyxl as op
import psycopg2

conn = psycopg2.connect(database='postgres', user='postgres', password='Vehfniby7', host='localhost', port=5432)
cursor = conn.cursor() 
conn.autocommit = True


data = op.load_workbook('data.xlsx', data_only=True)

sheets = data.sheetnames
main_sheets_values = None


def first_sheet():
    sheet_1 = sheets[0]
    sheet = data[sheet_1]
    sql_sheet = sheet_1.replace(" ", "_")
    print(sql_sheet)

    create_table_query = f"CREATE TABLE {sql_sheet} ("

    row_num = 2
    col_start = 2
    col_finish = 4

# Part1. Создание заголовка таблицы
    for row in sheet.iter_rows(min_row=row_num, max_row=row_num, min_col=col_start, max_col=col_finish,values_only=True):
            # Проверяем каждую строку на наличие данных            это лучше, это для вычисления первой строки(заголовков) 
        start_row = row
                # print(start_row) with None
        for cell_value in start_row:
                print(cell_value) # without None
                        # print(type(cell_value))
                column_name = str(cell_value).replace(" ", "_").replace("п\п", "id").replace("-", "_").replace("/", "_")
                        
                        # if type(cell_value) == str:
                        #     sql_zag_type == "VARCHAR(250)"    возможность реализации под опреденные типы данных
                        # if type(cell_value) == int:
                        #     sql_zag_type == "INTEGER"
                            # Приведение заголовков к нижнему регистру и замена пробелов на нижние подчеркивания
                create_table_query += f"{column_name} VARCHAR(250),"  # Задайте тип данных по умолчанию или подберите соответствующий вашим данным
        break
    create_table_query = create_table_query.rstrip(',') + ");"
    cursor.execute(create_table_query)
    print("База данных успешно создана")


# Part2. Заполнение созданной таблицы данными
    insert_values = []
    for row in sheet.iter_rows(min_row=row_num+1, min_col=col_start, max_col=col_finish, values_only=True):
        row_values = []
        for cell_value in row:
            row_values.append(str(cell_value).replace(" ", "_"))
        insert_values.append(tuple(row_values))
        
        # Проверяем, что список insert_values не пустой, перед выполнением запроса INSERT INTO
    if insert_values:
        for row_values in insert_values:
            insert_query = f"INSERT INTO {sql_sheet} VALUES %s;"
            cursor.execute(insert_query, (row_values,))


def second_sheet():
    sheet_1 = sheets[1]
    sheet = data[sheet_1]
    sql_sheet = sheet_1.replace(" ", "_")
    print(sql_sheet)

    create_table_query = f"CREATE TABLE {sql_sheet} ("

    row_num = 3
    col_start = 1
    col_finish = 2

# Part1. Создание заголовка таблицы
    for row in sheet.iter_rows(min_row=row_num, max_row=row_num, min_col=col_start, max_col=col_finish,values_only=True):
            # Проверяем каждую строку на наличие данных            это лучше, это для вычисления первой строки(заголовков) 
        start_row = row
                # print(start_row) with None
        for cell_value in start_row:
                print(cell_value) # without None
                        # print(type(cell_value))
                column_name = str(cell_value).replace(" ", "_").replace("п\п", "id").replace("-", "_").replace("/", "_")
                        
                        # if type(cell_value) == str:
                        #     sql_zag_type == "VARCHAR(250)"    возможность реализации под опреденные типы данных
                        # if type(cell_value) == int:
                        #     sql_zag_type == "INTEGER"
                            # Приведение заголовков к нижнему регистру и замена пробелов на нижние подчеркивания
                create_table_query += f"{column_name} VARCHAR(250),"  # Задайте тип данных по умолчанию или подберите соответствующий вашим данным
        break
    create_table_query = create_table_query.rstrip(',') + ");"
    cursor.execute(create_table_query)
    print("База данных успешно создана")


# Part2. Заполнение созданной таблицы данными
    insert_values = []
    for row in sheet.iter_rows(min_row=row_num+1, min_col=col_start, max_col=col_finish, values_only=True):
        row_values = []
        for cell_value in row:
            row_values.append(str(cell_value).replace(" ", "_"))
        insert_values.append(tuple(row_values))
        
        # Проверяем, что список insert_values не пустой, перед выполнением запроса INSERT INTO
    if insert_values:
        for row_values in insert_values:
            insert_query = f"INSERT INTO {sql_sheet} VALUES %s;"
            cursor.execute(insert_query, (row_values,))


def third_sheet():
    sheet_1 = sheets[2]   # Меняем тут страницу !!!!!!!!
    sheet = data[sheet_1]
    sql_sheet = sheet_1.replace(" ", "_").replace("-", "_").replace(".", "_")
    print(sql_sheet)

    sql_sheet = '"' + str(sql_sheet) + '"'

    print(sql_sheet)

    create_table_query = f"CREATE TABLE {sql_sheet} ("

    row_num = 2      # Меняем тут начальную строку !!!!!!!!
    col_start = 1    # Меняем тут начальный столбец !!!!!!!!
    col_finish = 28

# Part1. Создание заголовка таблицы
    for row in sheet.iter_rows(min_row=row_num, max_row=row_num, min_col=col_start, max_col=col_finish, values_only=True):
            # Проверяем каждую строку на наличие данных            это лучше, это для вычисления первой строки(заголовков) 
        start_row = row
                # print(start_row) with None
        for cell_value in start_row:
                print(cell_value) # without None
                column_name = str(cell_value).replace(" ", "_").replace("п\п", "id").replace("-", "_").replace("/", "_")
                column_name = '"' + column_name + '"'  # Оборачиваем имя столбца в кавычки
                create_table_query += f"{column_name} VARCHAR(250),"  # Задайте тип данных по умолчанию или подберите соответствующий вашим данным
        break
    create_table_query = create_table_query.rstrip(',') + ");"
    cursor.execute(create_table_query)
    print("База данных успешно создана")


# Part2. Заполнение созданной таблицы данными
    insert_values = []
    for row in sheet.iter_rows(min_row=row_num+1, min_col=col_start, max_col=col_finish, values_only=True):
        row_values = []
        for cell_value in row:
            row_values.append(str(cell_value).replace(" ", "_"))
        insert_values.append(tuple(row_values))
        
        # Проверяем, что список insert_values не пустой, перед выполнением запроса INSERT INTO

    # Проверяем, что список insert_values не пустой, перед выполнением запроса INSERT INTO
    if insert_values:
        for row_values in insert_values:
            insert_query = f"INSERT INTO {sql_sheet} VALUES %s;"
            cursor.execute(insert_query, (row_values,))

def fourth_sheet():
    sheet_1 = sheets[3]   # Меняем тут страницу !!!!!!!!
    sheet = data[sheet_1]
    sql_sheet = sheet_1.replace(" ", "_").replace("-", "_").replace(".", "_")
    print(sql_sheet)

    sql_sheet = '"' + str(sql_sheet) + '"'

    print(sql_sheet)

    create_table_query = f"CREATE TABLE {sql_sheet} ("

    row_num = 1      # Меняем тут начальную строку !!!!!!!!
    col_start = 7    # Меняем тут начальный столбец !!!!!!!!
    col_finish = 10

# Part1. Создание заголовка таблицы
    for row in sheet.iter_rows(min_row=row_num, max_row=row_num, min_col=col_start, max_col=col_finish, values_only=True):
            # Проверяем каждую строку на наличие данных            это лучше, это для вычисления первой строки(заголовков) 
        start_row = row
                # print(start_row) with None
        for cell_value in start_row:
                # print(cell_value) # without None
                column_name = str(cell_value).replace(" ", "_").replace("п\п", "id").replace("-", "_").replace("/", "_")
                column_name = '"' + column_name + '"'  # Оборачиваем имя столбца в кавычки
                create_table_query += f"{column_name} VARCHAR(250),"  # Задайте тип данных по умолчанию или подберите соответствующий вашим данным
        break
    create_table_query = create_table_query.rstrip(',') + ");"
    cursor.execute(create_table_query)
    print("База данных успешно создана!")


# Part2. Заполнение созданной таблицы данными
    insert_values = []
    for row in sheet.iter_rows(min_row=row_num+1, max_row=16, min_col=col_start, max_col=col_finish, values_only=True):
        row_values = []
        for cell_value in row:
            row_values.append(str(cell_value).replace(" ", "_"))
        insert_values.append(tuple(row_values))
        
        # Проверяем, что список insert_values не пустой, перед выполнением запроса INSERT INTO

    # Проверяем, что список insert_values не пустой, перед выполнением запроса INSERT INTO
    if insert_values:
        for row_values in insert_values:
            insert_query = f"INSERT INTO {sql_sheet} VALUES %s;"
            cursor.execute(insert_query, (row_values,))

    print("Данные в БД успешно вставлены!")

def fifth_sheet():
    sheet_1 = sheets[4]   # Меняем тут страницу !!!!!!!!
    sheet = data[sheet_1]
    sql_sheet = sheet_1.replace(" ", "_").replace("-", "_").replace(".", "_")
    print(sql_sheet)

    sql_sheet = '"' + str(sql_sheet) + '"'

    print(sql_sheet)

    create_table_query = f"CREATE TABLE {sql_sheet} ("

    row_num = 3      # Меняем тут начальную строку !!!!!!!!
    col_start = 1    # Меняем тут начальный столбец !!!!!!!!
    col_finish = 3

# Part1. Создание заголовка таблицы
    for row in sheet.iter_rows(min_row=row_num, max_row=row_num, min_col=col_start, max_col=col_finish, values_only=True):
            # Проверяем каждую строку на наличие данных            это лучше, это для вычисления первой строки(заголовков) 
        start_row = row
                # print(start_row) with None
        for cell_value in start_row:
                # print(cell_value) # without None
                column_name = str(cell_value).replace(" ", "_").replace("п\п", "id").replace("-", "_").replace("/", "_")
                column_name = '"' + column_name + '"'  # Оборачиваем имя столбца в кавычки
                create_table_query += f"{column_name} VARCHAR(250),"  # Задайте тип данных по умолчанию или подберите соответствующий вашим данным
        break
    create_table_query = create_table_query.rstrip(',') + ");"
    cursor.execute(create_table_query)
    print("База данных успешно создана!")


# Part2. Заполнение созданной таблицы данными
    insert_values = []
    for row in sheet.iter_rows(min_row=row_num+1, max_row=5, min_col=col_start, max_col=col_finish, values_only=True):
        row_values = []
        for cell_value in row:
            row_values.append(str(cell_value).replace(" ", "_"))
        insert_values.append(tuple(row_values))
        
        # Проверяем, что список insert_values не пустой, перед выполнением запроса INSERT INTO

    # Проверяем, что список insert_values не пустой, перед выполнением запроса INSERT INTO
    if insert_values:
        for row_values in insert_values:
            insert_query = f"INSERT INTO {sql_sheet} VALUES %s;"
            cursor.execute(insert_query, (row_values,))

    print("Данные в БД успешно вставлены!")

first_sheet()
second_sheet()
third_sheet()
fourth_sheet()
fifth_sheet()

cursor.close()
conn.close()